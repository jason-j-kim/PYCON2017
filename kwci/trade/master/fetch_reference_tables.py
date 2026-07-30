#!/usr/bin/env python3
"""
KWCI L1 — 기준 대응표(reference table) 수집·정규화

두 개의 UNSD 공개 대응표를 받아 공통 스키마 CSV로 정규화한다.

  1) HS2017 <-> HS2022 correlation
       용도: 2018~2025 수집 구간이 두 HS 판에 걸쳐 있어 전 구간을 HS2022로
            통일하지 않으면 2022년에 가짜 점프/붕괴가 생긴다.
  2) HS -> BEC rev.5 correspondence
       용도: S0 필터. 최종소비재(Consumption goods)만 후보로 남긴다.
            화장품 원료↔완제품, 원단↔의류가 여기서 자동 분리된다.

인증키 불필요(공개 자료). 단 UN 도메인이 막힌 네트워크에서는 --manual 경로를 쓴다.

UNSD는 대응표 경로·파일명 규칙이 일관되지 않으므로(같은 계열에 CPCv21_HS12 와
CPCv21_HS2017 이 공존) URL을 추측하지 않고 목록 페이지의 링크를 탐색한다.

사용법
------
  python fetch_reference_tables.py --list         # 탐색된 후보 링크만 점수순 출력
  python fetch_reference_tables.py                # 탐색 + 다운로드 + 정규화
  python fetch_reference_tables.py --normalize    # 이미 raw/에 받아둔 파일만 정규화

산출
----
  master/raw/                 원본 보존 (재현성)
  master/hs2017_hs2022.csv    hs2017, hs2022, relation
  master/hs_bec5.csv          hs, hs_edition, bec5, bec_class
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import sys
import zipfile
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import unquote, urljoin

try:
    import requests
except ImportError:
    sys.exit("requests가 필요합니다:  pip install requests openpyxl")

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"

# UNSD는 대응표 파일 경로·명명 규칙이 일관되지 않다
# (같은 계열인데 .../tables/CPC/CPCv21_HS12/ 와 .../tables/CPC/CPCv21_HS2017/ 이 공존).
# 그래서 URL을 추측하지 않고 목록 페이지에서 링크를 탐색한다.
DISCOVERY_PAGES = [
    "https://unstats.un.org/unsd/trade/classifications/correspondence-tables.asp",
    "https://unstats.un.org/unsd/classifications/Econ",
    "https://unstats.un.org/unsd/classifications/Econ/tables",
]

# 자동 탐색이 실패했을 때 사람이 직접 볼 곳. 순서대로 시도할 것.
FALLBACK_PAGES = [
    ("UNSD 대응표 목록",
     "https://unstats.un.org/unsd/trade/classifications/correspondence-tables.asp"),
    ("WCO 공식 HS2017-2022 correlation table",
     "https://www.wcoomd.org/en/topics/nomenclature/instrument-and-tools/"
     "hs-nomenclature-2022-edition/correlation-tables-hs-2017-2022.aspx"),
    ("World Bank WITS product concordance (HS 판간·HS-BEC)",
     "https://wits.worldbank.org/product_concordance.html"),
    ("UNSD BEC Rev.5 매뉴얼 (대응표 소재 안내 포함)",
     "https://unstats.un.org/unsd/trade/bec%20classification.htm"),
]

DATA_EXT = (".txt", ".csv", ".xlsx", ".xls", ".zip")

SOURCES: dict[str, dict] = {
    "hs2017_hs2022": {
        "label": "HS2017 <-> HS2022 correlation",
        # (필수 토큰군, 가점 토큰)
        "require": [("2017", "h17"), ("2022", "h22")],
        "bonus": ("hs", "correlation", "correspondence", "conversion"),
    },
    "hs_bec5": {
        "label": "HS -> BEC rev.5 correspondence",
        "require": [("bec",), ("hs",)],
        "bonus": ("5", "rev5", "rev.5", "correspondence", "conversion"),
    },
}

# --- 최종용도(end-use) 판정 -------------------------------------------------
#
# S0 필터는 "최종소비재만 남긴다"이므로 BEC 코드를 end-use로 환산해야 한다.
# 판정 경로는 셋이고, 위에서부터 우선한다.
#
#   (1) 파일에 end-use 열이 있으면 그것을 쓴다.
#       BEC rev.5는 end-use를 코드 자릿수가 아니라 별도 속성으로 두므로,
#       rev.5 대응표를 받았다면 반드시 이 경로로 가야 한다.
#   (2) 코드가 BEC rev.4 basic category 형태면 아래 표준 매핑을 쓴다.
#   (3) 둘 다 아니면 'unclassified'로 남기고 경고한다. 추정하지 않는다.
#
# rev.4 매핑은 UNSD가 문서화한 SNA end-use 구분이다.
BEC4_ENDUSE = {
    "capital":      {"41", "521"},
    "intermediate": {"111", "121", "21", "22", "31", "322", "42", "53"},
    "consumption":  {"112", "122", "51", "522", "61", "62", "63"},
    # 321(motor spirit)은 관례가 갈려 배정하지 않는다. K-도메인과 무관하다.
}
BEC4_LOOKUP = {c: cls for cls, codes in BEC4_ENDUSE.items() for c in codes}

ENDUSE_WORDS = {
    "consumption": "consumption", "consumer": "consumption",
    "intermediate": "intermediate",
    "capital": "capital",
}

TIMEOUT = 60
UA = {"User-Agent": "KWCI-L1/0.1 (reference table fetcher)"}


# ---------------------------------------------------------------- 다운로드

class _LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []   # (href, 링크 텍스트)
        self._href: str | None = None
        self._buf: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag == "a":
            self._href = dict(attrs).get("href")
            self._buf = []

    def handle_data(self, data):
        if self._href:
            self._buf.append(data)

    def handle_endtag(self, tag):
        if tag == "a" and self._href:
            self.links.append((self._href, " ".join(self._buf).strip()))
            self._href, self._buf = None, []


def discover() -> list[tuple[str, str]]:
    """목록 페이지들을 훑어 (절대 URL, 링크 텍스트) 목록을 모은다."""
    found: dict[str, str] = {}
    for page in DISCOVERY_PAGES:
        try:
            r = requests.get(page, headers=UA, timeout=TIMEOUT)
        except requests.RequestException as e:
            print(f"  x {type(e).__name__}  {page}")
            continue
        if r.status_code != 200:
            print(f"  x HTTP {r.status_code}  {page}")
            continue
        p = _LinkParser()
        p.feed(r.text)
        for href, text in p.links:
            if not href or href.startswith(("#", "mailto:", "javascript:")):
                continue
            found.setdefault(urljoin(r.url, href), text)
        print(f"  o {len(p.links):>4} 링크  {page}")
    return sorted(found.items())


def score(url: str, text: str, spec: dict) -> int:
    """링크가 찾는 대응표일 가능성을 점수화. 0 이하면 후보에서 제외."""
    hay = f"{unquote(url)} {text}".lower()

    # 필수 토큰군: 각 그룹에서 최소 하나는 나와야 한다.
    for group in spec["require"]:
        if not any(t in hay for t in group):
            return 0

    s = 10
    s += sum(3 for t in spec["bonus"] if t in hay)

    ext = Path(unquote(url).split("?")[0]).suffix.lower()
    if ext in DATA_EXT:
        s += 20                       # 기계가 읽을 수 있는 데이터 파일
    elif ext == ".pdf":
        s -= 15                       # 기술노트·매뉴얼은 대응표가 아니다
    elif ext in ("", ".asp", ".htm", ".html"):
        s += 2                        # 폴더/색인 페이지 — 한 단계 더 들어가야 함

    for bad in ("technical_note", "readme", "manual", "presentation", "note"):
        if bad in hay:
            s -= 12
    return s


def download(key: str, spec: dict, links: list[tuple[str, str]]) -> Path | None:
    ranked = sorted(
        ((score(u, t, spec), u, t) for u, t in links),
        key=lambda x: -x[0],
    )
    ranked = [r for r in ranked if r[0] > 0]
    if not ranked:
        print("    · 목록 페이지에서 후보 링크를 찾지 못했습니다")
        return None

    print(f"    · 후보 {len(ranked)}건, 상위부터 시도")
    RAW.mkdir(parents=True, exist_ok=True)

    for sc, url, text in ranked[:6]:
        label = text[:50] or Path(unquote(url)).name
        try:
            r = requests.get(url, headers=UA, timeout=TIMEOUT)
        except requests.RequestException as e:
            print(f"      x [{sc:>3}] {type(e).__name__}  {label}")
            continue
        if r.status_code != 200 or not r.content:
            print(f"      x [{sc:>3}] HTTP {r.status_code}  {label}")
            continue

        ext = Path(unquote(url).split("?")[0]).suffix.lower()
        if ext not in DATA_EXT:
            print(f"      · [{sc:>3}] 데이터 파일 아님(건너뜀)  {label}")
            continue

        dest = RAW / f"{key}{ext}"
        dest.write_bytes(r.content)
        print(f"      o [{sc:>3}] {len(r.content):,} bytes -> {dest.name}   {url}")
        return dest
    return None


# ---------------------------------------------------------------- 파싱

def read_rows(path: Path) -> list[list[str]]:
    """UNSD 배포 형식(txt/csv/zip/xlsx)을 행 리스트로 통일해 읽는다."""
    suffix = path.suffix.lower()

    if suffix == ".zip":
        with zipfile.ZipFile(path) as z:
            inner = next(
                (n for n in z.namelist()
                 if n.lower().endswith((".txt", ".csv", ".xlsx"))), None
            )
            if inner is None:
                raise ValueError(f"{path.name}: zip 안에 읽을 파일이 없습니다")
            tmp = RAW / Path(inner).name
            tmp.write_bytes(z.read(inner))
            return read_rows(tmp)

    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit("xlsx를 읽으려면:  pip install openpyxl")
        ws = load_workbook(path, read_only=True, data_only=True).active
        return [
            ["" if c is None else str(c).strip() for c in row]
            for row in ws.iter_rows(values_only=True)
        ]

    text = path.read_bytes().decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delim = max(",", ";", "\t", key=sample.count)
    return [
        [c.strip().strip('"') for c in row]
        for row in csv.reader(io.StringIO(text), delimiter=delim)
        if any(c.strip() for c in row)
    ]


HS_RE = re.compile(r"^\d{6}$")


def clean_hs(v: str) -> str | None:
    """HS 6단위로 정규화. 점·공백만 허용하고 6자리(또는 선행 0 소실된 5자리)만 인정.

    영문자가 섞인 값은 거부한다 — 이게 없으면 헤더 "HS 2017"이 숫자만 남아
    006자리로 zfill되어 HS코드 002017로 둔갑한다(실제로 겪은 버그).
    """
    v = (v or "").strip()
    if not v or re.search(r"[A-Za-z]", v):
        return None
    d = re.sub(r"[.\s\-]", "", v)
    if not d.isdigit():
        return None
    if len(d) == 5:
        d = "0" + d             # 엑셀이 삼킨 선행 0 복원 (010121 -> 10121)
    elif len(d) > 6:
        d = d[:6]               # HSK 10단위 등은 6단위로 절단
    return d if HS_RE.match(d) else None


def find_hs_columns(rows: list[list[str]]) -> tuple[int, int, int]:
    """헤더가 판마다 달라, 데이터 행에서 HS 코드처럼 보이는 열 2개를 찾는다."""
    scan = rows[: min(len(rows), 400)]
    width = max(len(r) for r in scan)
    score = [sum(1 for r in scan if i < len(r) and clean_hs(r[i])) for i in range(width)]
    ranked = sorted(range(width), key=lambda i: score[i], reverse=True)
    if len(ranked) < 2 or score[ranked[1]] == 0:
        raise ValueError("HS 코드 열을 2개 찾지 못했습니다 — 원본 형식을 확인하세요")
    a, b = sorted(ranked[:2])
    header_rows = 0                      # 선두 연속 비(非)데이터 행만 건너뛴다
    for r in rows[:5]:
        if len(r) > a and clean_hs(r[a]):
            break
        header_rows += 1
    return a, b, header_rows


def normalize_correlation(path: Path, out: Path) -> int:
    rows = read_rows(path)
    a, b, skip = find_hs_columns(rows)

    # 파일이 HS2022->HS2017 방향일 수도 있어 열 순서를 헤더로 판정한다.
    head = " ".join(rows[0]).lower() if rows else ""
    first_is_2017 = True
    if "2022" in head and "2017" in head:
        first_is_2017 = head.index("2022") > head.index("2017")

    pairs: list[tuple[str, str]] = []
    for r in rows[skip:]:
        if len(r) <= b:
            continue
        x, y = clean_hs(r[a]), clean_hs(r[b])
        if x and y:
            pairs.append((x, y) if first_is_2017 else (y, x))

    if not pairs:
        raise ValueError(f"{path.name}: 유효한 HS 쌍이 없습니다")

    # 관계 유형 판정: 1:N(분할)·N:1(통합)은 안분 규칙이 필요하므로 표시해 둔다.
    from collections import Counter
    fwd, bwd = Counter(x for x, _ in pairs), Counter(y for _, y in pairs)

    seen: set[tuple[str, str]] = set()
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["hs2017", "hs2022", "relation"])
        for x, y in pairs:
            if (x, y) in seen:
                continue
            seen.add((x, y))
            rel = "1:1"
            if fwd[x] > 1 and bwd[y] > 1:
                rel = "N:N"
            elif fwd[x] > 1:
                rel = "1:N"      # HS2017 하나가 HS2022 여럿으로 분할
            elif bwd[y] > 1:
                rel = "N:1"      # HS2017 여럿이 HS2022 하나로 통합
            w.writerow([x, y, rel])
    return len(seen)


BEC_RE = re.compile(r"^\d(\d|[A-Za-z])*$")


def normalize_bec(path: Path, out: Path) -> tuple[int, int]:
    rows = read_rows(path)
    scan = rows[: min(len(rows), 400)]
    width = max(len(r) for r in scan)

    hs_scores = [sum(1 for r in scan if i < len(r) and clean_hs(r[i])) for i in range(width)]
    hs_col = max(range(width), key=lambda i: hs_scores[i])
    if hs_scores[hs_col] == 0:
        raise ValueError("HS 코드 열을 찾지 못했습니다")

    def bec_score(i: int) -> int:
        if i == hs_col:
            return -1
        n = 0
        for r in scan:
            if i < len(r):
                v = r[i].strip()
                if v and BEC_RE.match(v) and len(v) <= 5 and not clean_hs(v):
                    n += 1
        return n

    bec_col = max(range(width), key=bec_score)
    if bec_score(bec_col) <= 0:
        raise ValueError("BEC 코드 열을 찾지 못했습니다 — 원본 형식을 확인하세요")

    # (1) end-use 열 탐색 — consumption/intermediate/capital 문자열이 있는 열.
    def enduse_score(i: int) -> int:
        if i in (hs_col, bec_col):
            return -1
        return sum(
            1 for r in scan
            if i < len(r) and any(w in r[i].lower() for w in ENDUSE_WORDS)
        )

    enduse_col = max(range(width), key=enduse_score)
    if enduse_score(enduse_col) <= 0:
        enduse_col = None

    edition = "HS2022" if "2022" in path.stem or "2022" in " ".join(rows[0] if rows else []) else "HS2017"

    def classify(bec: str, row: list[str]) -> str:
        if enduse_col is not None and enduse_col < len(row):
            low = row[enduse_col].lower()
            for word, cls in ENDUSE_WORDS.items():
                if word in low:
                    return cls
        return BEC4_LOOKUP.get(bec, "unclassified")

    rec: dict[str, tuple[str, str]] = {}
    for r in rows:
        if len(r) <= max(hs_col, bec_col):
            continue
        hs = clean_hs(r[hs_col])
        bec = r[bec_col].strip()
        if not hs or not bec or not BEC_RE.match(bec):
            continue
        rec[hs] = (bec, classify(bec, r))

    if not rec:
        raise ValueError(f"{path.name}: 유효한 HS-BEC 쌍이 없습니다")

    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["hs", "hs_edition", "bec", "bec_class", "enduse_source"])
        src = "file" if enduse_col is not None else "bec4_mapping"
        for hs in sorted(rec):
            bec, cls = rec[hs]
            w.writerow([hs, edition, bec, cls, src])

    consumption = sum(1 for _, c in rec.values() if c == "consumption")
    unknown = sum(1 for _, c in rec.values() if c == "unclassified")
    if unknown:
        share = unknown / len(rec)
        print(f"    ! end-use 미판정 {unknown:,}건 ({share:.0%}) — 출처={src}")
        if share > 0.2:
            print("      BEC rev.5 end-use 속성표를 별도로 받아 조인해야 합니다.")
    return len(rec), consumption


# ---------------------------------------------------------------- main

def manual_notice() -> None:
    pages = "\n".join(f"         - {n}\n           {u}" for n, u in FALLBACK_PAGES)
    print(f"""
  자동 탐색 실패. 수동 경로:

    1) 아래 페이지를 브라우저로 열어 대응표 파일을 내려받습니다
{pages}

    2) 받은 파일을 {RAW} 에 저장 (없으면 폴더 생성)
         hs2017_hs2022.<txt|csv|xlsx|zip>   HS2017 <-> HS2022 correlation
         hs_bec5.<txt|csv|xlsx|zip>         HS -> BEC 대응표

       파일명은 위 접두어만 맞으면 확장자는 무관합니다.

    3) python fetch_reference_tables.py --normalize

  --list 로 탐색된 링크 전체를 점수순으로 볼 수 있습니다.
  올바른 링크를 찾으셨다면 알려주시면 스크립트에 고정하겠습니다.
""")


def find_raw(key: str) -> Path | None:
    if not RAW.exists():
        return None
    return next((p for p in sorted(RAW.iterdir())
                 if p.is_file() and p.stem == key), None)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--list", action="store_true",
                    help="탐색된 링크를 점수순으로 출력만 (다운로드 안 함)")
    ap.add_argument("--normalize", action="store_true", help="raw/의 기존 파일만 정규화")
    args = ap.parse_args()

    links: list[tuple[str, str]] = []
    if not args.normalize:
        print("[목록 페이지 탐색]")
        links = discover()
        if not links:
            print("\n  목록 페이지에 접근하지 못했습니다.")
            manual_notice()
            return 1

    if args.list:
        for key, spec in SOURCES.items():
            print(f"\n[{spec['label']}] 후보")
            ranked = sorted(((score(u, t, spec), u, t) for u, t in links),
                            key=lambda x: -x[0])
            hits = [r for r in ranked if r[0] > 0][:15]
            if not hits:
                print("  (없음)")
            for sc, u, t in hits:
                print(f"  [{sc:>3}] {t[:55]:<55} {u}")
        return 0

    failed = []
    for key, spec in SOURCES.items():
        print(f"\n[{spec['label']}]")

        path = find_raw(key) if args.normalize else None
        if path:
            print(f"    · 기존 파일 사용: {path.relative_to(HERE)}")
        elif not args.normalize:
            path = download(key, spec, links) or find_raw(key)

        if not path:
            print("    ! 원본 없음")
            failed.append(key)
            continue

        try:
            if key == "hs2017_hs2022":
                n = normalize_correlation(path, HERE / "hs2017_hs2022.csv")
                print(f"    = hs2017_hs2022.csv  {n:,} 매핑")
            else:
                n, cons = normalize_bec(path, HERE / "hs_bec5.csv")
                print(f"    = hs_bec5.csv  {n:,} 품목  (소비재 {cons:,} — S0 통과 후보)")
        except Exception as e:
            print(f"    ! 정규화 실패: {e}")
            failed.append(key)

    if failed:
        manual_notice()
        return 1

    print("\n완료. S0(소비재 필터)·HS2022 통일을 돌릴 준비가 됐습니다.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
